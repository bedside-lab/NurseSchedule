"""
Core scheduling engine: solve -> validate -> build_output.
This is the same rule set developed and validated for 제2중환자실's monthly
roster, refactored so a GUI (or anything else) can call it with a plain
config dict instead of editing a script's CONFIG block by hand.

config keys (all required unless noted):
    input_file            path to the source .xlsx (requests pre-filled)
    output_file            path to write the finished .xlsx to
    sheet_name              default 'Sheet1'
    header_row_start        0-indexed pandas row where nurse data begins (default 3)
    name_col                0-indexed pandas column with the nurse's name (default 1)
    team_col                0-indexed pandas column with the nurse's team (default 2)
    first_date_col          0-indexed pandas column of the first history day (default 3)
    n_history_days          trailing days of the previous month shown for context
    n_target_days           number of days in the month being scheduled
    dayonly_nurses          list[str] of nurses who are Day-shift only this month
    special_code_map        dict like {'S': 'O', 'Y': 'O', 'F': 'D'}
    staff_target            dict {'D': 5, 'E': 5, 'N': 4}
    staff_bounds            dict {'D': (4,6), 'E': (4,6), 'N': (3,5)}
    weekly_min_off          default 2
    night_immediate_off_days default 2
    night_min_gap_days      default 6
    max_consec_workdays     default 5
    n_range_max             default 1
    e_range_max             default 5
    solver_time_limit_sec   default 280
"""

import copy
import json

import openpyxl
import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from ortools.sat.python import cp_model

CATS = ['D', 'E', 'N', 'O']
YELLOW = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
ORANGE = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')


def _cfg(config, key, default=None):
    return config.get(key, default)


def _conv_hist(v, special_map):
    if not isinstance(v, str):
        return None
    if v in special_map:
        return special_map[v]
    if v in CATS:
        return v
    return None


def load_nurses(config):
    df = pd.read_excel(config['input_file'], sheet_name=_cfg(config, 'sheet_name', 'Sheet1'), header=None)
    header_row_start = _cfg(config, 'header_row_start', 3)
    name_col = _cfg(config, 'name_col', 1)
    team_col = _cfg(config, 'team_col', 2)
    first_date_col = _cfg(config, 'first_date_col', 3)
    n_hist = config['n_history_days']
    n_target = config['n_target_days']
    special_map = config.get('special_code_map', {'S': 'O', 'Y': 'O', 'F': 'D'})
    dayonly = set(config.get('dayonly_nurses', []))

    hist_end = first_date_col + n_hist
    target_end = hist_end + n_target

    nurses = []
    for i in range(header_row_start, df.shape[0]):
        name = str(df.iat[i, name_col]).strip()
        if name in ('', 'nan', 'None'):
            continue
        team = str(df.iat[i, team_col]).strip()
        hist = [df.iat[i, j] for j in range(first_date_col, hist_end)]
        sep = [df.iat[i, j] for j in range(hist_end, target_end)]
        nurses.append(dict(row=i, name=name, team=team, hist=hist, sep=list(sep)))
    for n in nurses:
        n['hist_conv'] = [_conv_hist(v, special_map) for v in n['hist']]
        n['pregnant'] = n['name'] in dayonly
        n['histg'] = {(-n_hist + k): n['hist_conv'][k] for k in range(n_hist)}
    return nurses, df


def _compute_hist_forbidden_window(histg, night_min_gap_days):
    gs = sorted(histg.keys())
    vals = [histg[g] for g in gs]
    n = len(vals)
    forbidden = set()
    k = 0
    while k < n:
        if vals[k] == 'N':
            k2 = k
            while k2 < n and vals[k2] == 'N':
                k2 += 1
            if k2 < n:
                end_g = gs[k2 - 1]
                for kk in range(1, night_min_gap_days + 1):
                    g = end_g + kk
                    if g >= 0:
                        forbidden.add(g)
            k = k2
        else:
            k += 1
    return forbidden


def solve(config, log=print):
    nurses, _ = load_nurses(config)
    NDAYS = config['n_target_days']
    NUM = len(nurses)
    teams = sorted(set(n['team'] for n in nurses))

    night_min_gap_days = _cfg(config, 'night_min_gap_days', 6)
    night_immediate_off_days = _cfg(config, 'night_immediate_off_days', 2)
    weekly_min_off = _cfg(config, 'weekly_min_off', 2)
    max_consec_workdays = _cfg(config, 'max_consec_workdays', 5)
    n_range_max = _cfg(config, 'n_range_max', 1)
    e_range_max = _cfg(config, 'e_range_max', 5)
    staff_target = _cfg(config, 'staff_target', {'D': 5, 'E': 5, 'N': 4})
    staff_bounds = _cfg(config, 'staff_bounds', {'D': (4, 6), 'E': (4, 6), 'N': (3, 5)})
    n_hist = config['n_history_days']

    for n in nurses:
        n['hist_forbidden'] = _compute_hist_forbidden_window(n['histg'], night_min_gap_days)

    model = cp_model.CpModel()
    x = {}
    for i, n in enumerate(nurses):
        for d in range(NDAYS):
            for c in CATS:
                x[i, d, c] = model.NewBoolVar(f'x_{i}_{d}_{c}')
            model.Add(sum(x[i, d, c] for c in CATS) == 1)

    for i, n in enumerate(nurses):
        if n['pregnant']:
            for d in range(NDAYS):
                model.Add(x[i, d, 'E'] == 0)
                model.Add(x[i, d, 'N'] == 0)

    # a nurse's own REQUESTED Off day can never be preceded by a Night shift
    for n in nurses:
        n['requested_off_days'] = {d for d in range(NDAYS) if n['sep'][d] == 'O'}
    unavoidable = [n['name'] for n in nurses
                   if 0 in n['requested_off_days'] and n['histg'][-1] == 'N']
    if unavoidable:
        log(f'참고: 이전달 마지막날 이미 나이트라 "신청 오프 앞 나이트 금지" 규칙을 지킬 수 없는 경우: {unavoidable}')

    fixed_count = 0
    overridden = []
    for i, n in enumerate(nurses):
        for d in range(NDAYS):
            v = n['sep'][d]
            if isinstance(v, str) and v in CATS:
                if v == 'N' and d in n['hist_forbidden']:
                    overridden.append((n['name'], d + 1, 'night-interval'))
                    continue
                if v == 'N' and (d + 1) in n['requested_off_days']:
                    overridden.append((n['name'], d + 1, 'before-requested-off'))
                    continue
                model.Add(x[i, d, v] == 1)
                fixed_count += 1
    log(f'고정된 신청 근무: {fixed_count}칸')
    if overridden:
        log(f'규칙 충돌로 조정된 신청: {overridden}')

    isN = lambda i, d: x[i, d, 'N']
    isO = lambda i, d: x[i, d, 'O']
    isE = lambda i, d: x[i, d, 'E']
    isD = lambda i, d: x[i, d, 'D']

    def isNg(i, g): return isN(i, g) if g >= 0 else (1 if nurses[i]['histg'][g] == 'N' else 0)
    def isOg(i, g): return isO(i, g) if g >= 0 else (1 if nurses[i]['histg'][g] == 'O' else 0)
    def isEg(i, g): return isE(i, g) if g >= 0 else (1 if nurses[i]['histg'][g] == 'E' else 0)
    def isDg(i, g): return isD(i, g) if g >= 0 else (1 if nurses[i]['histg'][g] == 'D' else 0)

    for i in range(NUM):
        for g in range(-1, NDAYS - 1):
            model.Add(isEg(i, g) + isDg(i, g + 1) <= 1)

    for i in range(NUM):
        for g in range(-1, NDAYS - 1):
            model.Add(isNg(i, g) <= isNg(i, g + 1) + isOg(i, g + 1))

    for i in range(NUM):
        for g in range(-night_immediate_off_days, NDAYS - night_immediate_off_days):
            model.Add(isNg(i, g) + isOg(i, g + 1) - isOg(i, g + 2) <= 1)

    for i in range(NUM):
        for g in range(-n_hist, NDAYS - 1):
            for k in range(1, night_min_gap_days + 1):
                gk = g + k
                if gk > NDAYS - 1 or gk < 0:
                    continue
                model.Add(isNg(i, g) - isNg(i, g + 1) + isNg(i, gk) <= 1)

    # a nurse's own REQUESTED Off day can never be preceded by a Night shift
    for i, n in enumerate(nurses):
        for d in n['requested_off_days']:
            g_before = d - 1
            if g_before >= 0:
                model.Add(isN(i, g_before) == 0)

    forced_first_night = []
    for i, n in enumerate(nurses):
        if n['pregnant']:
            continue
        hg = n['histg']
        if hg[-1] == 'N' and hg[-2] != 'N':
            model.Add(x[i, 0, 'N'] == 1)
            forced_first_night.append(n['name'])
    if forced_first_night:
        log(f'전월 마지막날 시작된 나이트를 이어서 배정: {forced_first_night}')

    week_starts = list(range(-1, NDAYS - 6, 7))
    for i, n in enumerate(nurses):
        for ws in week_starts:
            days = list(range(ws, ws + 7))
            if all(g < 0 for g in days):
                continue
            model.Add(sum(isOg(i, g) for g in days) >= weekly_min_off)

    for i in range(NUM):
        for g in range(-max_consec_workdays, NDAYS - max_consec_workdays):
            model.Add(sum(isOg(i, g + k) for k in range(max_consec_workdays + 1)) >= 1)

    dev_terms = []
    for d in range(NDAYS):
        Dcnt = sum(isD(i, d) for i in range(NUM))
        Ecnt = sum(isE(i, d) for i in range(NUM))
        Ncnt = sum(isN(i, d) for i in range(NUM))
        dD = model.NewIntVar(0, NUM, f'dD_{d}')
        dE = model.NewIntVar(0, NUM, f'dE_{d}')
        dN = model.NewIntVar(0, NUM, f'dN_{d}')
        model.AddAbsEquality(dD, Dcnt - staff_target['D'])
        model.AddAbsEquality(dE, Ecnt - staff_target['E'])
        model.AddAbsEquality(dN, Ncnt - staff_target['N'])
        model.Add(Dcnt >= staff_bounds['D'][0]); model.Add(Dcnt <= staff_bounds['D'][1])
        model.Add(Ecnt >= staff_bounds['E'][0]); model.Add(Ecnt <= staff_bounds['E'][1])
        model.Add(Ncnt >= staff_bounds['N'][0]); model.Add(Ncnt <= staff_bounds['N'][1])
        dev_terms.append(20 * dD + 15 * dE + 35 * dN)

    team_excess_terms = []
    for d in range(NDAYS):
        for c in ['D', 'E', 'N']:
            for t in teams:
                members = [i for i, n in enumerate(nurses) if n['team'] == t]
                cnt = sum(x[i, d, c] for i in members)
                excess = model.NewIntVar(0, len(members), f'exc_{d}_{c}_{t}')
                model.Add(excess >= cnt - 1)
                team_excess_terms.append(excess)

    nonpreg = [i for i, n in enumerate(nurses) if not n['pregnant']]

    Ntot = {i: sum(isN(i, d) for d in range(NDAYS)) for i in nonpreg}
    Nmax = model.NewIntVar(0, NDAYS, 'Nmax'); Nmin = model.NewIntVar(0, NDAYS, 'Nmin')
    model.AddMaxEquality(Nmax, [Ntot[i] for i in nonpreg])
    model.AddMinEquality(Nmin, [Ntot[i] for i in nonpreg])
    Nrange = Nmax - Nmin
    model.Add(Nrange <= n_range_max)

    Etot = {i: sum(isE(i, d) for d in range(NDAYS)) for i in nonpreg}
    Emax = model.NewIntVar(0, NDAYS, 'Emax'); Emin = model.NewIntVar(0, NDAYS, 'Emin')
    model.AddMaxEquality(Emax, [Etot[i] for i in nonpreg])
    model.AddMinEquality(Emin, [Etot[i] for i in nonpreg])
    Erange = Emax - Emin
    model.Add(Erange <= e_range_max)

    Ototl = {i: sum(isO(i, d) for d in range(NDAYS)) for i in range(NUM)}
    Omax = model.NewIntVar(0, NDAYS, 'Omax'); Omin = model.NewIntVar(0, NDAYS, 'Omin')
    model.AddMaxEquality(Omax, [Ototl[i] for i in range(NUM)])
    model.AddMinEquality(Omin, [Ototl[i] for i in range(NUM)])
    Orange = Omax - Omin

    isolated_terms = []
    for i in range(NUM):
        for d in range(1, NDAYS - 1):
            a = isO(i, d - 1); b = isO(i, d); c = isO(i, d + 1)
            zo = model.NewBoolVar(f'iso_o_{i}_{d}')
            model.Add(zo <= 1 - a); model.Add(zo <= b); model.Add(zo <= 1 - c)
            model.Add(zo >= (1 - a) + b + (1 - c) - 2)
            isolated_terms.append(zo)
            zw = model.NewBoolVar(f'iso_w_{i}_{d}')
            model.Add(zw <= a); model.Add(zw <= 1 - b); model.Add(zw <= c)
            model.Add(zw >= a + (1 - b) + c - 2)
            isolated_terms.append(zw)

    single_night_terms = []
    for i in range(NUM):
        for d in range(0, NDAYS - 1):
            prev = isNg(i, d - 1); cur = isN(i, d); nxt = isNg(i, d + 1)
            z = model.NewBoolVar(f'single_n_{i}_{d}')
            model.Add(z <= 1 - prev); model.Add(z <= cur); model.Add(z <= 1 - nxt)
            model.Add(z >= (1 - prev) + cur + (1 - nxt) - 2)
            single_night_terms.append(z)

    long_streak_terms = []
    for i in range(NUM):
        for g in range(-3, NDAYS - 3):
            vals = [isNg(i, g + k) for k in range(4)]
            z = model.NewBoolVar(f'long_n_{i}_{g}')
            for v in vals:
                model.Add(z <= v)
            model.Add(z >= sum(vals) - 3)
            long_streak_terms.append(z)

    objective = (
        sum(dev_terms)
        + 6 * sum(team_excess_terms)
        + 25 * Nrange
        + 8 * Erange
        + 12 * Orange
        + 3 * sum(isolated_terms)
        + 45 * sum(single_night_terms)
        + 30 * sum(long_streak_terms)
    )
    model.Minimize(objective)

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = _cfg(config, 'solver_time_limit_sec', 280)
    solver.parameters.num_search_workers = 8
    solver.parameters.log_search_progress = False
    log('스케줄 계산 중... (최대 몇 분 정도 걸릴 수 있습니다)')
    status = solver.Solve(model)
    log(f'solver 상태: {solver.StatusName(status)}')
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError('조건을 만족하는 근무표를 찾지 못했습니다 (INFEASIBLE). 신청 근무나 규칙 값을 확인해주세요.')

    result = []
    for i, n in enumerate(nurses):
        row = []
        for d in range(NDAYS):
            for c in CATS:
                if solver.Value(x[i, d, c]) == 1:
                    row.append(c)
                    break
        result.append(row)

    solution = {
        'names': [n['name'] for n in nurses],
        'teams': [n['team'] for n in nurses],
        'schedule': result,
        'overridden_requests': overridden,
        'forced_first_night': forced_first_night,
        'night_min_gap_days_used': night_min_gap_days,
    }
    return solution


def solve_with_fallback(config, log=print):
    """Try the preferred night-interval gap first; if that makes the whole
    schedule infeasible, retry with the fallback gap (never the other way
    around -- the longer gap is always preferred when it works)."""
    primary = _cfg(config, 'night_min_gap_days', 6)
    fallback = _cfg(config, 'night_min_gap_days_fallback', None)
    try:
        cfg1 = dict(config, night_min_gap_days=primary)
        return solve(cfg1, log=log)
    except RuntimeError as e:
        if fallback is None or fallback == primary:
            raise
        log(f'나이트 인터벌 {primary}일로는 조건을 만족하는 근무표를 찾지 못했습니다. '
            f'{fallback}일 인터벌로 다시 시도합니다...')
        cfg2 = dict(config, night_min_gap_days=fallback)
        return solve(cfg2, log=log)


def validate(config, solution, log=print):
    nurses, _ = load_nurses(config)
    names, sched = solution['names'], solution['schedule']
    overridden_set = {(item[0], item[1]) for item in solution.get('overridden_requests', [])}
    n_hist = config['n_history_days']
    n_target = config['n_target_days']
    dayonly = set(config.get('dayonly_nurses', []))
    weekly_min_off = _cfg(config, 'weekly_min_off', 2)
    # validate against whichever gap the solver actually used (it may have
    # fallen back from the preferred value -- see solve_with_fallback)
    night_min_gap_days = solution.get('night_min_gap_days_used', _cfg(config, 'night_min_gap_days', 6))
    max_consec_workdays = _cfg(config, 'max_consec_workdays', 5)

    errors = []
    for n, row in zip(nurses, sched):
        name = n['name']
        hist_conv = n['hist_conv']
        full = hist_conv + row

        def g2idx(g):
            return g + n_hist

        for d in range(n_target):
            req = n['sep'][d]
            if isinstance(req, str) and req in CATS:
                if row[d] != req and (name, d + 1) not in overridden_set:
                    errors.append(f'{name} day{d + 1}: requested {req} but got {row[d]}')

        if name in dayonly:
            for d in range(n_target):
                if row[d] not in ('D', 'O'):
                    errors.append(f'{name} day{d + 1}: day-only nurse but got {row[d]}')

        for g in range(-1, n_target):
            idx1, idx2 = g2idx(g), g2idx(g + 1)
            if idx2 >= len(full):
                continue
            if full[idx1] == 'E' and full[idx2] == 'D':
                errors.append(f'{name} g{g}: Evening followed by Day')
            if full[idx1] == 'N' and full[idx2] not in ('N', 'O'):
                errors.append(f'{name} g{g}: Night followed by {full[idx2]}')
            if full[idx1] == 'N' and full[idx2] == 'O':
                idx3 = g2idx(g + 2)
                if idx3 < len(full) and full[idx3] != 'O':
                    errors.append(f'{name} g{g}: fewer than mandatory offs after night streak')

        for g in range(-n_hist, n_target - 1):
            idx1, idx2 = g2idx(g), g2idx(g + 1)
            if full[idx1] == 'N' and full[idx2] != 'N':
                for k in range(1, night_min_gap_days + 1):
                    gk = g + k
                    idxk = g2idx(gk)
                    if idxk >= len(full):
                        continue
                    if full[idxk] == 'N':
                        errors.append(f'{name}: night streak ended g{g}, N again at g{gk}')

        week_starts = list(range(-1, n_target - 6, 7))
        for ws in week_starts:
            idxs = [g2idx(g) for g in range(ws, ws + 7) if 0 <= g2idx(g) < len(full)]
            offs = sum(1 for k in idxs if full[k] == 'O')
            if offs < weekly_min_off:
                errors.append(f'{name} week g{ws}: only {offs} offs')

        if name not in dayonly:
            hg = n['histg']
            if hg[-1] == 'N' and hg[-2] != 'N' and row[0] != 'N':
                errors.append(f'{name}: boundary lone night not extended')

        run = 0
        for k in range(len(full)):
            if full[k] != 'O':
                run += 1
                if run > max_consec_workdays:
                    errors.append(f'{name}: >{max_consec_workdays} consecutive workdays ending idx{k}')
            else:
                run = 0

        for d in range(n_target):
            if n['sep'][d] == 'O':
                idx_before = g2idx(d - 1)
                if 0 <= idx_before < len(full) and full[idx_before] == 'N':
                    errors.append(f'{name}: Night immediately precedes their requested off on day{d + 1}')

    log(f'검증 결과: 오류 {len(errors)}건')
    for e in errors[:50]:
        log('  ' + e)
    return errors


def build_output(config, solution, log=print):
    input_file = config['input_file']
    output_file = config['output_file']
    sheet_name = _cfg(config, 'sheet_name', 'Sheet1')
    header_row_start = _cfg(config, 'header_row_start', 3)
    name_col = _cfg(config, 'name_col', 1)
    first_date_col = _cfg(config, 'first_date_col', 3)
    n_hist = config['n_history_days']
    n_target = config['n_target_days']
    label_cols = first_date_col  # columns before the date columns (1-indexed count)

    try:
        wb = openpyxl.load_workbook(input_file)
    except Exception:
        log('원본 파일을 바로 열 수 없어 LibreOffice로 복구를 시도합니다...')
        import subprocess
        import os
        outdir = os.path.join(os.path.dirname(output_file) or '.', '_repaired')
        os.makedirs(outdir, exist_ok=True)
        subprocess.run(
            ['soffice', '--headless', '--convert-to', 'xlsx:Calc MS Excel 2007 XML', '--outdir', outdir, input_file],
            check=True, capture_output=True,
        )
        repaired_path = os.path.join(outdir, os.path.splitext(os.path.basename(input_file))[0] + '.xlsx')
        wb = openpyxl.load_workbook(repaired_path)
    ws = wb[sheet_name]

    names, sched = solution['names'], solution['schedule']
    REASON_TEXT = {
        'night-interval': '나이트 인터벌(최소 간격) 규칙과 충돌',
        'before-requested-off': '본인이 신청한 다음날 오프 바로 앞에 나이트가 오지 않도록 하는 규칙과 충돌',
    }
    overridden = {}
    for item in solution.get('overridden_requests', []):
        nm, d = item[0], item[1]
        overridden[(nm, d)] = item[2] if len(item) > 2 else None

    df = pd.read_excel(input_file, sheet_name=sheet_name, header=None)

    first_date_col_xl = first_date_col + 1
    target_first_col_xl = first_date_col_xl + n_hist
    target_last_col_xl = target_first_col_xl + n_target - 1
    first_row_xl = header_row_start + 1

    last_row_xl = first_row_xl
    filled_blanks = 0
    requested_marked = 0
    name_idx = 0
    for i in range(header_row_start, df.shape[0]):
        name = str(df.iat[i, name_col]).strip()
        if name in ('', 'nan', 'None'):
            continue
        row_xl = i + 1
        last_row_xl = row_xl
        for d in range(n_target):
            col_xl = target_first_col_xl + d
            orig = df.iat[i, first_date_col + n_hist + d]
            was_requested = isinstance(orig, str) and orig in CATS
            cell = ws.cell(row=row_xl, column=col_xl)
            cell.value = sched[name_idx][d]
            if was_requested:
                requested_marked += 1
                key = (name, d + 1)
                if key in overridden:
                    reason = REASON_TEXT.get(overridden[key], overridden[key] or '규칙 충돌')
                    cell.fill = ORANGE
                    cell.comment = Comment(
                        f'원래 신청: {orig} ({reason}로 다른 근무로 조정됨)', '근무표 자동배정'
                    )
                else:
                    cell.fill = YELLOW
            else:
                filled_blanks += 1
        name_idx += 1

    header_style_cell = ws.cell(row=2, column=3)
    data_style_cell = ws.cell(row=first_row_xl, column=first_date_col_xl)

    def apply_header_style(cell):
        cell.font = copy.copy(header_style_cell.font)
        cell.alignment = copy.copy(header_style_cell.alignment)
        cell.fill = copy.copy(header_style_cell.fill)
        cell.border = copy.copy(header_style_cell.border)

    def apply_data_style(cell):
        cell.font = copy.copy(data_style_cell.font)
        cell.alignment = copy.copy(data_style_cell.alignment)

    right_col_start = target_last_col_xl + 1
    right_cols = {right_col_start + k: label for k, label in enumerate(['D', 'E', 'N', 'O'])}
    for col, label in right_cols.items():
        letter = get_column_letter(col)
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        top_cell = ws.cell(row=2, column=col)
        top_cell.value = label
        apply_header_style(top_cell)
        apply_header_style(ws.cell(row=3, column=col))
        ws.column_dimensions[letter].width = 6

    target_first_letter = get_column_letter(target_first_col_xl)
    target_last_letter = get_column_letter(target_last_col_xl)
    for r in range(first_row_xl, last_row_xl + 1):
        for col, label in right_cols.items():
            cell = ws.cell(row=r, column=col)
            cell.value = f'=COUNTIF({target_first_letter}{r}:{target_last_letter}{r},"{label}")'
            apply_data_style(cell)

    bottom_row_start = last_row_xl + 1
    bottom_rows = {bottom_row_start + k: label for k, label in enumerate(['D', 'E', 'N'])}
    for row, label in bottom_rows.items():
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=label_cols)
        label_cell = ws.cell(row=row, column=1)
        label_cell.value = f'{label} 인원수'
        for c in range(1, label_cols + 1):
            apply_header_style(ws.cell(row=row, column=c))
        for col in range(first_date_col_xl, target_last_col_xl + 1):
            letter = get_column_letter(col)
            cell = ws.cell(row=row, column=col)
            cell.value = f'=COUNTIF({letter}{first_row_xl}:{letter}{last_row_xl},"{label}")'
            apply_data_style(cell)
        for col in right_cols:
            apply_header_style(ws.cell(row=row, column=col))

    legend_row = bottom_row_start + len(bottom_rows) + 1
    ws.cell(row=legend_row, column=1).value = (
        '노란색 = 기존 신청 근무 반영 / 주황색 = 신청했으나 규칙 충돌로 조정됨(셀 코멘트 참고)'
    )

    wb.save(output_file)
    log(f'신청 반영 {requested_marked}칸, 새로 채운 칸 {filled_blanks}개')
    log(f'완성 파일 저장: {output_file}')


def run_pipeline(config, log=print):
    """Solve -> validate -> build output, in one call. Returns (solution, errors)."""
    solution = solve_with_fallback(config, log=log)
    errors = validate(config, solution, log=log)
    build_output(config, solution, log=log)
    return solution, errors
